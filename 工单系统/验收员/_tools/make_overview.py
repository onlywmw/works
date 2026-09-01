# -*- coding: utf-8 -*-
"""make_overview.py — 证据链总览表生成：扫描日期证据目录 → xlsx + md
用法: python make_overview.py <日期:YYYY-MM-DD> [输出路径前缀]
产出: <前缀>_总览.xlsx + <前缀>_总览.md（默认前缀 = 数据目录/<日期>/真机验收总览_<日期>）
扫描结构: <日期>/<场景>/*.png（每个 png 配同名 .ocr.txt）
"""
import os, sys, glob

try:
    import openpyxl
except ImportError:
    sys.exit("需要 openpyxl: python -m pip install openpyxl")

BASE = r"C:\Users\Administrator\Desktop\MOV\工单流转中心\验收员"
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

day = sys.argv[1] if len(sys.argv) > 1 else None
daydir = os.path.join(BASE, day) if day else max(
    (d for d in os.listdir(BASE) if os.path.isdir(os.path.join(BASE, d)) and d[0].isdigit()),
    default=None,
)
if not daydir or not os.path.isdir(daydir):
    sys.exit("找不到日期目录: %s" % daydir)

rows = []  # (场景, 证据名, png, ocr文本)
for scene in sorted(os.listdir(daydir)):
    sd = os.path.join(daydir, scene)
    if not os.path.isdir(sd):
        continue
    for png in sorted(glob.glob(os.path.join(sd, "*.png"))):
        name = os.path.splitext(os.path.basename(png))[0]
        txt = png + ".ocr.txt"
        text = ""
        if os.path.isfile(txt):
            with open(txt, "r", encoding="utf-8") as f:
                text = f.read().strip()
        rows.append((scene, name, png, text))

# ---- xlsx ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "证据总览"
ws.append(["序号", "场景", "证据名", "截图文件", "OCR 文本", "备注/结论"])
for i, (scene, name, png, text) in enumerate(rows, 1):
    first_line = text.split("\n")[0] if text else "(无OCR)"
    ws.append([i, scene, name, os.path.basename(png), first_line, ""])
# 全文本放备注列太长，增加第二 sheet 存全文
ws2 = wb.create_sheet("OCR全文")
ws2.append(["序号", "场景", "证据名", "OCR 全文"])
for i, (scene, name, png, text) in enumerate(rows, 1):
    ws2.append([i, scene, name, text])
widths = [6, 18, 22, 24, 60, 20]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
out_xlsx = os.path.join(BASE, "%s_总览.xlsx" % day)
wb.save(out_xlsx)

# ---- md ----
out_md = os.path.join(BASE, "%s_总览.md" % day)
lines = ["# 真机验收证据总览（%s）" % day, ""]
lines.append("| 序号 | 场景 | 证据名 | OCR 识别文本（首行） |")
lines.append("|---|---|---|---|")
for i, (scene, name, png, text) in enumerate(rows, 1):
    fl = (text.split("\n")[0] if text else "(无OCR)").replace("|", "\\|")
    lines.append("| %d | %s | %s | %s |" % (i, scene, name, fl))
lines.append("")
with open(out_md, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print("✅ 总览表已生成:")
print("  " + out_xlsx)
print("  " + out_md)
print("共 %d 条证据" % len(rows))
