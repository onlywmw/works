#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""审验.py —— SYS-01 阶段五 P2 反伪审验工具（工单系统内部工具 · 零 APP 影响）

定位：审验员 AI 逐一审验工单证据链的工具。机器只出 flag，人判结论。

接口：
    python 审验.py --ticket <工单号> [--json]   # 审验单工单证据链 + integrity_flags + integrity_review + 交付联动
    python 审验.py --list                        # 列出有证据的工单
    python 审验.py --dir <证据目录> [--json]     # 对指定证据目录直接审验（构造边界用例用）
    python 审验.py --manifest <manifest.json>    # 独立验证 evidence_manifest（P0-2 绑定 + 层B逐条）
    python 审验.py --coverage <交付报告.md>      # P1-1 能力护栏自动核验（NONE禁合/PARTIAL缺裁决打回/FULL过）
    python 审验.py --ticket-file <派单文本.md>   # P1-2 坑位库查询字段核对（缺失/否无说明 → 退回 flag）
    python 审验.py --verify-hash <branch> <hash> [--repo <仓库>]  # SYS-02 E2 hash 一致性自动闸（存在性 + origin/main 祖先校验）
    python 审验.py --verify-hash-self-test [--repo <仓库>]        # SYS-02 E2 回归自测（U-49 fixture 重放 9fd39b6→REJECT / 2a13dcd→OK）

依据：
    - 方案：设计师\方案设计\改造计划_工单流转中心防伪抗倒退_2026-09-01.md §P2-1（v3.1）
    - 交付绑定：工单系统\交付绑定规范.md（evidence_manifest / manifest_sha §四）
    - 能力护栏：验收员\文档\真机验收标准与方式方法.md §1.5（coverage_status 三档）
    - 知识库：设计师\知识库\README.md（出单前查询强制字段）
    - 证据链规范：审验员\证据链规范.md（journal/截图/命令/时间一致性 四环节）
    - E2 一致性：设计师\派单\SYS-02_阶段一_派单_2026-09-02.md（SYS-02 E2 hash 一致性自动闸）

红线：integrity_flags/能力护栏 flag 绝不自动等价"不通过"；最终判定由人工完成（integrity_review）。
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import os
import re
import subprocess
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent  # 工单系统根

# 证据根目录（同一工单证据只允许存在于"产出该证据的角色"目录下，见 README 红线）
EVIDENCE_ROOTS = [
    ROOT / "验收员" / "证据数据",
    ROOT / "程序员",
    ROOT / "检查证据",
    ROOT / "审验员" / "证据数据",
]

KNOWN_PRODUCERS = ("验收员", "程序员", "审验员", "设计师", "检查证据")

# UPG-86：路径嵌注释检测词表（path 应为纯路径；说明性内容须移 note 字段）
_PATH_NOTE_MARKERS = re.compile(r"[（）｜§]|——|sha256=")
# UPG-86：合法 sha256 形态（64 位十六进制）
_SHA256_RE = re.compile(r"^[0-9a-fA-F]{64}$")


def _dir_sha256(dirpath: Path) -> str:
    """目录聚合 sha256（UPG-86 口径：sorted 文件名+内容 依次摘要）——与交付报告声明口径一致。"""
    h = hashlib.sha256()
    for f in sorted(os.listdir(dirpath)):
        fp = os.path.join(dirpath, f)
        if os.path.isfile(fp):
            h.update(f.encode("utf-8"))
            h.update(Path(fp).read_bytes())
    return h.hexdigest()


def _has_missing_declare(ev: dict) -> bool:
    """显式 missing 声明：sha256=null 且 note 字段含 missing 标注（如实标注不造假——STD-UPG-86-v1）。"""
    if ev.get("sha256") is not None:
        return False
    note = str(ev.get("note", "")) + str(ev.get("missing_declare", ""))
    return "missing" in note.lower() or "missing" in note

# 强断言词（报告声称"已支持/已实现"类）—— semantic_vagueness 启发式
STRONG_CLAIM_RE = re.compile(r"已支持|已实现|已接入|已完成|已修复|可正常|全绿|验收通过|达成")
# 弱断言词（只查"存在/非空"类）—— benchmark_overfit 启发式
WEAK_ASSERT_RE = re.compile(r"存在|非空|!= ?null|== ?0|> ?0|不等于|不为空|无异常")


def _norm(s: str) -> str:
    """规范化工单号/目录名用于匹配：UPG-28 ↔ UPG28 ↔ 2026-08-30\\UPG28。"""
    return re.sub(r"[\W_]+", "", s).lower()


def json_dumps(o) -> str:
    return json.dumps(o, ensure_ascii=False, indent=2)


# ---------------------------------------------------------------- 时间戳解析

def parse_ts_value(v):
    """把 journal 行 / manifest 里的时间值转成 epoch 秒；无法解析返回 None。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        if x > 1e12:
            x /= 1000.0  # 毫秒 → 秒
        if 1e8 < x < 1e11:  # 秒级 epoch（1973~5138）
            return x
        return None
    s = str(v).strip()
    try:
        d = datetime.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return d.timestamp()
    except ValueError:
        pass
    try:
        x = float(s)
        if x > 1e12:
            x /= 1000.0
        if 1e8 < x < 1e11:
            return x
        return None
    except ValueError:
        return None


def _find_ts_field(obj, depth=0):
    if depth > 4 or not isinstance(obj, dict):
        return None
    for k in ("timestamp", "ts", "time", "t", "created_at", "date", "event_time", "recorded_at"):
        if k in obj and isinstance(obj[k], (int, float, str)):
            return obj[k]
    for v in obj.values():
        r = _find_ts_field(v, depth + 1)
        if r is not None:
            return r
    return None


def ts_from_filename(name: str):
    """从截图文件名提取时间戳。返回 (epoch_sec, 精确到时分秒?)。仅日期不算精确。"""
    # 紧凑格式（P2 补正则，修已知坑#5）：YYYYMMDD_HHMMSS / YYYYMMDD-HHMMSS / YYYYMMDD_HH-MM-SS
    m = re.search(
        r"(?P<y>\d{4})(?P<mo>\d{2})(?P<d>\d{2})[-_.](?P<H>\d{2})(?P<M>\d{2})(?P<S>\d{2})",
        name,
    )
    if not m:
        m = re.search(
            r"(?P<y>\d{4})(?P<mo>\d{2})(?P<d>\d{2})[-_.](?P<H>\d{2})[-_.](?P<M>\d{2})[-_.](?P<S>\d{2})",
            name,
        )
    if m:
        try:
            d = datetime.datetime(int(m.group("y")), int(m.group("mo")), int(m.group("d")),
                                  int(m.group("H")), int(m.group("M")), int(m.group("S")))
            return d.timestamp(), True
        except ValueError:
            pass
    # 带分隔符格式：YYYY-MM-DD HH:MM:SS / YYYY-MM-DD_HH:MM（既有）
    m = re.search(
        r"(?P<y>\d{4})[-_.](?P<mo>\d{2})[-_.](?P<d>\d{2})"
        r"[T _-]*(?P<H>\d{1,2})[:_.-](?P<M>\d{2})[:_.-]?(?P<S>\d{2})?",
        name,
    )
    if m:
        try:
            H, M = int(m.group("H")), int(m.group("M"))
            S = int(m.group("S")) if m.group("S") else 0
            d = datetime.datetime(int(m.group("y")), int(m.group("mo")), int(m.group("d")), H, M, S)
            return d.timestamp(), True
        except ValueError:
            pass
    m2 = re.search(r"(?P<y>\d{4})[-_.](?P<mo>\d{2})[-_.](?P<d>\d{2})", name)
    if m2:
        try:
            d = datetime.datetime(int(m2.group("y")), int(m2.group("mo")), int(m2.group("d")))
            return d.timestamp(), False
        except ValueError:
            pass
    return None, False


# ---------------------------------------------------------------- 证据收集

def collect_evidence(ev_dir: Path):
    """扫描一个证据目录，分类：journal(jsonl)/截图/命令与记录(md,txt,xml,log)。"""
    journal, screenshots, commands = [], [], []
    if not ev_dir.exists():
        return journal, screenshots, commands
    for p in sorted(ev_dir.rglob("*")):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        if ext == ".jsonl":
            journal.append(p)
        elif ext in (".png", ".jpg", ".jpeg"):
            screenshots.append(p)
        elif ext in (".md", ".txt", ".xml", ".log", ".dump", ".html"):
            commands.append(p)
    return journal, screenshots, commands


def journal_tool_calls(journal_files):
    """统计 journal 工具调用链：tool/call 或 tool_call，附最早/最晚时间戳。"""
    calls = 0
    tses = []
    for jf in journal_files:
        try:
            with open(jf, encoding="utf-8", errors="replace") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    if "tool/call" in line or "tool_call" in line:
                        calls += 1
                    try:
                        obj = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    tv = _find_ts_field(obj)
                    e = parse_ts_value(tv)
                    if e is not None:
                        tses.append(e)
        except OSError:
            continue
    earliest = min(tses) if tses else None
    latest = max(tses) if tses else None
    return calls, earliest, latest


def _fmt_ts(e):
    if e is None:
        return None
    return datetime.datetime.fromtimestamp(e).strftime("%Y-%m-%d %H:%M:%S")


def screenshot_times(screenshots):
    """每张截图的 (文件名, 精确时间戳或None, mtime)。"""
    out = []
    for p in screenshots:
        t, precise = ts_from_filename(p.name)
        out.append((p.name, t if precise else None, p.stat().st_mtime))
    return out


def chain_check(journal, screenshots, commands):
    """证据链四环节对账。"""
    jcalls, j_earliest, j_latest = journal_tool_calls(journal)
    times = screenshot_times(screenshots)
    time_flags, time_problems = [], []
    # 时序：截图(精确时间) vs journal 最早调用
    if j_earliest is not None:
        early_screens = [
            (name, t) for name, t, _ in times if t is not None and t < j_earliest
        ]
        if early_screens:
            time_flags.append("截图早于journal最早调用")
            time_problems.append(
                f"截图 {early_screens[0][0]}（{_fmt_ts(early_screens[0][1])}）早于 journal 最早工具调用（{_fmt_ts(j_earliest)}）——疑似时序不一致"
            )
    failure_hits = []
    for c in commands:
        try:
            txt = c.read_text(encoding="utf-8", errors="replace")
            if re.search(r"BUILD FAILED|FAILED|Exception|not found|不存在|失败", txt):
                failure_hits.append(c.name)
        except OSError:
            continue
    return {
        "journal_tool_calls": jcalls,
        "journal_time_range": [_fmt_ts(j_earliest), _fmt_ts(j_latest)],
        "screenshots": [name for name, _, _ in times],
        "commands": [c.name for c in commands],
        "time_consistency": {
            "flag": "异常" if time_flags else "正常",
            "problems": time_problems,
        },
        "failure_mark_in_text": failure_hits,  # 注意：文本可能只是描述性提到 FAILED，需人工复核
    }


def find_evidence_dirs(ticket):
    """按工单号在全部证据根目录找匹配目录。返回 [Path]。"""
    n = _norm(ticket)
    found = []
    for root in EVIDENCE_ROOTS:
        if not root.exists():
            continue
        for p in sorted(root.rglob("*")):
            if p.is_dir() and n in _norm(p.name):
                found.append(p)
    return found


# ---------------------------------------------------------------- 工单表

def load_ticket_table():
    """读工单表（升级工单表 sheet），返回 {工单号: {delivery_id, note}}；失败返回 (None, err)。"""
    try:
        import openpyxl
    except ImportError:
        return None, "缺少依赖 openpyxl（pip install openpyxl），无法读工单表"
    path = ROOT / "工单表.xlsx"
    if not path.exists():
        return None, f"工单表不存在: {path}"
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hdr = {c: (ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)}
        # 定位列
        col_num = next((c for c, v in hdr.items() if v == "工单号"), 1)
        col_del = next((c for c, v in hdr.items() if v == "delivery_id"), None)
        table = {}
        for r in range(2, ws.max_row + 1):
            tno = ws.cell(row=r, column=col_num).value
            if tno is None:
                continue
            dval = ws.cell(row=r, column=col_del).value if col_del else None
            table[str(tno).strip()] = {"delivery_id": str(dval) if dval else ""}
        return table, None
    except Exception as e:  # 防读坏崩溃
        return None, f"工单表读取失败: {e}"


# ---------------------------------------------------------------- P0-2 绑定

def _canon_manifest(manifest_list):
    """evidence_manifest 数组规范化 JSON 字节（键序/无空白/UTF-8）。"""
    return json.dumps(
        manifest_list, sort_keys=True, ensure_ascii=False, separators=(",", ":")
    ).encode("utf-8")


def _sha256_hex(b):
    return hashlib.sha256(b).hexdigest()


def verify_manifest(manifest_path):
    """独立验证 evidence_manifest：结构 + manifest_sha 重算比对 + 层B逐条（D 项）。"""
    res = {"manifest": str(manifest_path), "ok": False, "detail": {}}
    try:
        data = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
    except Exception as e:
        res["detail"]["error"] = f"manifest 读取/解析失败: {e}"
        return res
    mlist = data.get("evidence_manifest")
    if not isinstance(mlist, list):
        res["detail"]["error"] = "缺少 evidence_manifest 数组"
        return res
    problems = []
    for i, ev in enumerate(mlist, 1):
        for field in ("evidence_id", "path", "sha256", "producer", "created_at"):
            if field not in ev or not str(ev.get(field, "")).strip():
                # UPG-86：sha256 显式 null + missing 声明 = 如实标注（不造假），不算缺字段
                if field == "sha256" and ev.get("sha256") is None and _has_missing_declare(ev):
                    continue
                problems.append(f"EVID-{i} 缺字段 {field}")
    # UPG-86 检测 A：路径嵌注释（path 应为纯路径——说明性内容须移 note 字段）
    for i, ev in enumerate(mlist, 1):
        p = str(ev.get("path", ""))
        if _PATH_NOTE_MARKERS.search(p):
            problems.append(f"EVID-{i} 路径嵌注释：path 含说明性内容（（）/｜/§/——/sha256=），须移 note 字段后裸串化")
    # UPG-86 检测 B：缺 sha256（空串/缺字段/非 64hex 且无显式 missing 声明 → 应填未填，红）
    for i, ev in enumerate(mlist, 1):
        sha = ev.get("sha256")
        if sha is None:
            if not _has_missing_declare(ev):
                problems.append(f"EVID-{i} 缺 sha256 且无 missing 声明（如实标注须 sha256=null + note 含 missing）")
            continue
        sv = str(sha).strip()
        if not sv or not _SHA256_RE.match(sv):
            problems.append(f"EVID-{i} sha256 非法（空串或非 64 位十六进制）：{sv[:20]}")
    # manifest_sha 重算
    recomputed = _sha256_hex(_canon_manifest(mlist))
    bound = data.get("evidence_manifest_sha")
    match = None
    # UPG-86 检测 C：绑定值未写入（manifest 文件须自带 evidence_manifest_sha=清单内容重算值，可重算对账）
    if not bound:
        problems.append("绑定值未写入：manifest 文件缺 evidence_manifest_sha 字段（清单内容 canonical sha256，可重算对账）")
    else:
        match = (str(bound).strip().lower() == recomputed)
        if not match:
            problems.append("evidence_manifest_sha 与重算值不一致（清单内容已变或声明过期）")
    # 层B逐条
    integ = []
    for ev in mlist:
        p = ROOT / str(ev.get("path", ""))
        exists = p.exists()
        is_dir = exists and p.is_dir()
        declared_missing = _has_missing_declare(ev) or (not exists and "missing" in str(ev.get("note", "")).lower())
        hash_ok = None
        if exists and ev.get("sha256"):
            try:
                actual = _dir_sha256(p) if is_dir else _sha256_hex(p.read_bytes())
                hash_ok = (actual == str(ev["sha256"]).lower())
            except OSError:
                hash_ok = False
        integ.append({
            "evidence_id": ev.get("evidence_id"),
            "path": str(ev.get("path")),
            "exists": exists,
            "is_dir": is_dir,
            "hash_matches": hash_ok,
            "producer": ev.get("producer"),
            "producer_known": str(ev.get("producer", "")) in KNOWN_PRODUCERS,
            "missing_declared": declared_missing,
        })
        if not exists and not declared_missing:
            problems.append(f"{ev.get('evidence_id')} 路径不存在")
    res.update({
        "delivery_id": data.get("delivery_id"),
        "standard_id": data.get("standard_id"),
        "manifest_sha_bound": bound,
        "manifest_sha_recomputed": recomputed,
        "match": match,
        "evidence_count": len(mlist),
        "evidence_integrity": integ,
        "problems": problems,
        "ok": (not problems),
    })
    return res


def delivery_binding(ticket, table):
    """从工单表读该行 delivery_id 绑定值（P0-2 对接）。"""
    row = (table or {}).get(ticket)
    if row is None:
        return {"found_row": False, "note": f"工单表无 {ticket} 行"}
    dval = row.get("delivery_id", "")
    if not dval or dval.startswith("—"):
        return {"found_row": True, "delivery_id": dval or None, "bound": False,
                "note": "未绑定 delivery_id（改造前交付），跳过 hash 比对"}
    return {"found_row": True, "delivery_id": dval, "bound": True,
            "note": "已绑定 delivery_id，需按交付报告「交付绑定」节 / manifest 文件重算比对"}


# ---------------------------------------------------------------- SYS-02 E2 hash 一致性自动闸

# 主仓库（项目配置.md 第二节）：code_commit_sha 校验对象；--repo 可覆盖（测试/扩展用）
MAIN_REPO = r"C:\Users\Administrator\0027-mov"


def _run_git(repo, *args):
    """在指定仓库执行 git 命令（只读）。返回 (ok, out)；ok=False 时 out 为错误全文。"""
    try:
        p = subprocess.run(
            ["git", "-C", repo, *args],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
    except FileNotFoundError:
        return False, "git 命令不可用（未安装/不在 PATH）"
    except Exception as e:  # 防御仓库路径异常
        return False, f"git 执行异常: {e}"
    if p.returncode != 0:
        return False, (p.stderr or p.stdout or "").strip()
    return True, (p.stdout or "").strip()


def verify_hash(branch, reported_hash, repo=None):
    """E2 hash 一致性自动闸（SYS-02 阶段一）。

    在目标仓库（默认主仓库 0027-mov）校验交付报告的 code_commit_sha：
      ① git cat-file -t 校验存在且为 commit → 不存在 REJECT missing；短 hash 前缀模糊 REJECT ambiguous
      ② git merge-base --is-ancestor <hash> origin/main → 不在祖先链 REJECT not-ancestor
      ③ 通过 → HASH_OK
    机器只出 flag，登记/放行由人裁决。REJECT 提示用 `git log --oneline <branch>` 取当前真 hash。
    """
    repo = repo or MAIN_REPO
    res = {
        "mode": "verify-hash",
        "repo": repo,
        "branch": branch,
        "reported_hash": reported_hash,
        "status": None,
        "reason": None,
        "signal": "",
    }
    if not os.path.isdir(repo):
        res.update(status="HASH_REJECT", reason="missing", signal=f"目标仓库不存在: {repo}")
        return res
    # ① 存在性 + 对象类型
    ok, out = _run_git(repo, "cat-file", "-t", reported_hash)
    if not ok:
        if "ambiguous" in out.lower():
            res.update(
                status="HASH_REJECT",
                reason="ambiguous",
                signal=f"短 hash `{reported_hash}` 前缀无法唯一解析（ambiguous）——用完整 40 位 hash",
            )
        else:
            res.update(
                status="HASH_REJECT",
                reason="missing",
                signal=f"`{reported_hash}` 不存在（git cat-file -t 失败: {out}）——提示: `git log --oneline {branch}` 取当前真 hash",
            )
        return res
    if out.strip() != "commit":
        res.update(
            status="HASH_REJECT",
            reason="ambiguous",
            signal=f"`{reported_hash}` 对象类型={out.strip()}，非 commit——交付绑定需 commit hash",
        )
        return res
    # ② 祖先链校验（防旧 hash / 未合内容冒充）
    ok, out = _run_git(repo, "merge-base", "--is-ancestor", reported_hash, "origin/main")
    if not ok:
        if "origin/main" in out:
            res.update(
                status="FLAG",
                signal=f"本地无 origin/main（{out}）——先 `git fetch origin` 再校验（机器出 flag，人裁决）",
            )
        else:
            res.update(
                status="HASH_REJECT",
                reason="not-ancestor",
                signal=f"`{reported_hash}` 不在 origin/main 祖先链（{out}）——疑似 rebase 重写/未合内容冒充；提示: `git log --oneline {branch}` 取当前真 hash",
            )
        return res
    res.update(
        status="HASH_OK",
        signal=f"`{reported_hash}` 存在且在 origin/main 祖先链 → hash 一致性通过（可登记交付）",
    )
    return res


VERIFY_HASH_FIXTURES = [
    {
        "branch": "feat/upg49",
        "reported_hash": "9fd39b6",
        "expect": "HASH_REJECT",
        "reason": "missing",
        "desc": "UPG-49 交付报告 hash 9fd39b6（rebase 重写后已不存在）→ REJECT missing",
    },
    {
        "branch": "feat/upg49",
        "reported_hash": "2a13dcd",
        "expect": "HASH_OK",
        "reason": None,
        "desc": "UPG-49 分支第 3 commit 2a13dcd（在 origin/main 祖先链）→ HASH_OK",
    },
]


def verify_hash_self_test(repo=None):
    """E2 回归自测（派单回归锚）：重放 U-49 案例，连真实主仓库（只读）跑 fixture。"""
    repo = repo or MAIN_REPO
    cases = []
    for fx in VERIFY_HASH_FIXTURES:
        r = verify_hash(fx["branch"], fx["reported_hash"], repo)
        expect = f"{fx['expect']} {fx['reason'] or ''}".strip()
        actual = f"{r['status']} {r['reason'] or ''}".strip()
        cases.append(
            {
                "case": fx["desc"],
                "expected": expect,
                "actual": actual,
                "passed": actual == expect,
                "signal": r["signal"],
            }
        )
    return {
        "mode": "verify-hash-self-test",
        "repo": repo,
        "cases": cases,
        "ok": all(c["passed"] for c in cases),
    }


# ---------------------------------------------------------------- UPG-86 manifest 治理（三类失效检测自测）

def manifest_self_test():
    """UPG-86 亲杀锚自测：三坏案（路径嵌注释/缺 sha256/绑定值未写入）全红 + 好案绿。fixture 临时目录生成即弃。"""
    import tempfile

    def write_manifest(evs, bound):
        arr = json.dumps(evs, ensure_ascii=False)
        return "{\n  \"delivery_id\": \"DEL-TEST\",\n  \"evidence_manifest\": " + arr + ",\n  \"evidence_manifest_sha\": " + (
            json.dumps(bound) if bound else "null") + "\n}"

    def good_ev(**kw):
        base = {"evidence_id": "E-1", "path": "p.bin", "sha256": "a" * 64, "producer": "程序员", "created_at": "2026-09-03", "note": ""}
        base.update(kw)
        return base

    cases = []
    with tempfile.TemporaryDirectory() as td:
        # 好案素材：真实文件（绝对路径——verify_manifest 以 ROOT 基准拼接，绝对路径 pathlib 语义直取）+ 真实 sha256
        bin_path = Path(td) / "p.bin"
        bin_path.write_bytes(b"upg86-fixture")
        real_sha = _sha256_hex(bin_path.read_bytes())
        good = [good_ev(path=str(bin_path), sha256=real_sha)]
        good_manifest = Path(td) / "good.json"
        good_manifest.write_text(write_manifest(good, _sha256_hex(_canon_manifest(good))), encoding="utf-8")

        # 坏案①：路径嵌注释
        bad1 = [good_ev(path="0027-mov/ACCEPTANCE_LOG.md §P22-R1")]
        bad1_f = Path(td) / "bad1.json"
        bad1_f.write_text(write_manifest(bad1, _sha256_hex(_canon_manifest(bad1))), encoding="utf-8")

        # 坏案②：缺 sha256（空串无 missing 声明）
        bad2 = [good_ev(sha256="")]
        bad2_f = Path(td) / "bad2.json"
        bad2_f.write_text(write_manifest(bad2, _sha256_hex(_canon_manifest(bad2))), encoding="utf-8")

        # 坏案③：绑定值未写入（bound=null）
        bad3_f = Path(td) / "bad3.json"
        bad3_f.write_text(write_manifest(good, None), encoding="utf-8")

        for name, f, expect_ok in (
            ("好案（规范 manifest）", good_manifest, True),
            ("坏案①路径嵌注释", bad1_f, False),
            ("坏案②缺 sha256", bad2_f, False),
            ("坏案③绑定值未写入", bad3_f, False),
        ):
            r = verify_manifest(f)
            cases.append({
                "case": name,
                "expected_ok": expect_ok,
                "actual_ok": r["ok"],
                "passed": r["ok"] == expect_ok,
                "problems": r.get("problems", [])[:3],
            })
    return {
        "mode": "manifest-self-test",
        "cases": cases,
        "ok": all(c["passed"] for c in cases),
    }


# ---------------------------------------------------------------- P1-1 能力护栏（P2 自动化）

def _extract_section(text, title):
    """提取 md 中 `## ...title...` 到下一个 ## 标题之间的正文（标题行不跨行，正文跨行）。"""
    m = re.search(rf"^#+\s*[^\n]*{title}[^\n]*\n(.*?)(?=^#+\s|\Z)", text, re.M | re.S)
    return m.group(1) if m else None


def verify_coverage(report_path):
    """P1-1 能力护栏自动核验（机器只出 flag）：读交付报告「能力护栏」节 → 判 coverage_status。

    NONE → 禁合 main flag；PARTIAL 缺 coverage_decision 六字段或 decided_by≠设计师 → 打回 flag；
    PARTIAL 六字段齐 → 通过 flag；FULL → 通过。机器只出 flag，放行由验收员/设计师裁决。
    """
    res = {"report": str(report_path), "ok": False, "coverage_status": None,
           "flags": {}, "signal": "", "problems": []}
    try:
        text = Path(report_path).read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        res["problems"].append(f"报告读取失败: {e}")
        return res
    section = _extract_section(text, "能力护栏")
    if section is None:
        # 是否共享面变更：含共享面清单或四类关键词
        if re.search(r"MainActivity\s*注册表|工具面|协议\s*[·/]?\s*接口定义|全局数据结构|共享面影响清单", text):
            res["flags"]["block_merge"] = True
            res["problems"].append("共享面变更但无「能力护栏」节——打回（缺 coverage_status 不得放行）")
            res["signal"] = "共享面变更 + 无能力护栏节 → 打回"
        else:
            res["signal"] = "无能力护栏节（非共享面变更，N/A）"
        res["conclusion"] = "需人工裁决（机器只出 flag）"
        return res
    m = re.search(r"coverage_status\s*[:：]\s*(\w+)", section)
    if not m:
        res["flags"]["missing_status"] = True
        res["problems"].append("能力护栏节缺 coverage_status 字段——打回")
        res["signal"] = "缺 coverage_status → 打回"
        res["conclusion"] = "需人工裁决（机器只出 flag）"
        return res
    status = m.group(1).upper()
    res["coverage_status"] = status
    if status == "NONE":
        res["flags"]["block_merge"] = True
        res["signal"] = "coverage_status=NONE → 禁止合 main（明确禁合，非建议谨慎）"
        res["problems"].append("NONE 禁合 main")
    elif status == "FULL":
        res["flags"]["pass"] = True
        res["signal"] = "coverage_status=FULL → 直接过能力护栏"
    elif status == "PARTIAL":
        fields = {}
        for f in ("uncovered", "risk", "merge_decision", "reason", "decided_by", "decided_at"):
            fm = re.search(rf"^\s*{f}\s*[:：]\s*(\S[^\n]*)", section, re.M)
            fields[f] = fm.group(1).strip() if fm else None
        missing = [f for f, v in fields.items() if not v]
        if missing:
            res["flags"]["reject"] = True
            res["signal"] = f"PARTIAL 缺 coverage_decision 字段 {missing} → 打回（缺设计裁决记录不得放行）"
            res["problems"].append(f"缺字段: {missing}")
        elif "设计师" not in (fields.get("decided_by") or ""):
            res["flags"]["reject"] = True
            res["signal"] = f"PARTIAL 裁决人 {fields.get('decided_by')}≠设计师 → 打回"
            res["problems"].append("decided_by 必须为设计师")
        else:
            res["flags"]["pass"] = True
            res["signal"] = "PARTIAL + 完整设计师裁决（uncovered/risk/merge_decision/reason/decided_by=设计师/decided_at 齐）→ 通过"
            res["coverage_decision"] = fields
    else:
        res["flags"]["unknown"] = True
        res["problems"].append(f"未知 coverage_status: {status}")
        res["signal"] = f"coverage_status={status} 无法识别 → 待人工"
    res["ok"] = (not res["problems"])
    res["conclusion"] = "需人工裁决（机器只出 flag）"
    return res


def check_ticket_query_field(ticket_file):
    """P1-2 坑位库查询字段核对（P2 自动化）：读派单文本「已查坑位库/复用件库」字段。

    缺失或「否」无说明 → 退回 flag；已填（是/否+说明）→ 通过 flag（人工核对命中条目）。
    """
    res = {"ticket_file": str(ticket_file), "ok": False, "flags": {}, "signal": "", "problems": []}
    try:
        text = Path(ticket_file).read_text(encoding="utf-8", errors="replace")
    except OSError as e:
        res["problems"].append(f"派单文本读取失败: {e}")
        res["conclusion"] = "需人工裁决（机器只出 flag）"
        return res
    m = re.search(r"已查坑位库/复用件库\s*[:：]\s*(\S[^\n]*)", text)
    if not m:
        res["flags"]["reject"] = True
        res["signal"] = "派单文本缺「已查坑位库/复用件库」字段 → 退回（出单不合规）"
        res["problems"].append("缺「已查坑位库/复用件库」字段")
        res["conclusion"] = "需人工裁决（机器只出 flag）"
        return res
    val = m.group(1).strip()
    if re.fullmatch(r"否[。．.]?", val):
        res["flags"]["reject"] = True
        res["signal"] = "已查库字段填「否」且无说明 → 退回（须写明为何不查）"
        res["problems"].append("填「否」无说明")
    else:
        res["flags"]["pass"] = True
        res["signal"] = f"已查库字段已填: 「{val[:60]}」 → 通过（人工核对命中条目）"
    res["ok"] = (not res["problems"])
    res["conclusion"] = "需人工裁决（机器只出 flag）"
    return res


# ---------------------------------------------------------------- 四分类 flag

def compute_integrity_flags(chain, commands, screenshots):
    """层A 行为欺骗四分类（启发式机器 flag，配 signal 供人工裁决）。"""
    flags = {k: {"flag": False, "signal": ""} for k in
             ("original_rerun", "benchmark_overfit", "semantic_vagueness", "time_inconsistency")}

    # ④ time_inconsistency：截图(精确时间) 早于 journal 最早调用
    if chain["time_consistency"]["problems"]:
        flags["time_inconsistency"]["flag"] = True
        flags["time_inconsistency"]["signal"] = chain["time_consistency"]["problems"][0]

    # ③ semantic_vagueness：报告强断言 vs 执行证据缺失
    texts = []
    for c in commands:
        try:
            texts.append(c.read_text(encoding="utf-8", errors="replace"))
        except OSError:
            pass
    all_text = "\n".join(texts)
    claims = STRONG_CLAIM_RE.findall(all_text) if all_text else []
    exec_evidence = chain["journal_tool_calls"] > 0 or bool(chain["screenshots"])
    if claims and not exec_evidence:
        flags["semantic_vagueness"]["flag"] = True
        flags["semantic_vagueness"]["signal"] = (
            f"记录文本含强断言 {len(claims)} 处（{claims[:3]}…），但无 journal 工具调用且无截图——"
            "「声称已支持 X」与执行证据缺失，需人工比对报告声明 vs 代码实现"
        )
    else:
        flags["semantic_vagueness"]["signal"] = (
            f"检查路径：比对交付报告强断言（已支持/已实现…，本目录 {len(claims)} 处）与代码/证据是否一致；"
            "脚本只能标记不能代判，见 integrity_review"
        )

    # ② benchmark_overfit：断言弱（只查存在）而覆盖率高 → 提示人工
    weak = WEAK_ASSERT_RE.findall(all_text) if all_text else []
    if weak and len(weak) >= 5 and chain["journal_tool_calls"] == 0:
        flags["benchmark_overfit"]["flag"] = True
        flags["benchmark_overfit"]["signal"] = (
            f"文本含弱断言 {len(weak)} 处（存在/非空/==0 类）且无 journal——疑似只查存在不查语义，"
            "需人工检查测试是否引用基准/产物本身"
        )
    else:
        flags["benchmark_overfit"]["signal"] = (
            f"检查路径：静态看测试断言是否引用基准/产物本身；本目录弱断言词 {len(weak)} 处"
        )

    # ① original_rerun：截图时间戳批量重复 + 无 journal 变化 → 提示
    times = screenshot_times(screenshots)
    precise = [t for _, t, _ in times if t is not None]
    if len(precise) >= 3 and len(set(precise)) == 1 and chain["journal_tool_calls"] == 0:
        flags["original_rerun"]["flag"] = True
        flags["original_rerun"]["signal"] = "多张截图时间戳完全相同且无 journal——疑似原样重跑（仅改时间戳/窗口）"
    else:
        flags["original_rerun"]["signal"] = (
            "检查路径：git diff 交付分支 vs 工单表登记 commit；仅时间戳/窗口变化=原样重跑嫌疑"
        )
    return flags


def evidence_integrity_layer(chain, screenshots, commands, manifest_result=None):
    """层B 证据对象真实性六项（独立层，不混入四分类）。screenshots/commands 均为 Path 列表。"""
    files = screenshots + commands
    files_exist = bool(files)
    hash_matches = None
    if manifest_result is not None:
        hash_matches = manifest_result.get("match")
    now = datetime.datetime.now().timestamp()
    ts_plausible = all(p.stat().st_mtime <= now + 5 for p in files)
    producer_known = any(any(k in str(p) for k in KNOWN_PRODUCERS) for p in files)
    chain_intact = (
        chain["journal_tool_calls"] > 0
        and bool(chain["screenshots"])
        and bool(chain["commands"])
    )
    return {
        "file_exists": files_exist,
        "hash_matches": hash_matches,  # None = 无 manifest 绑定，N/A
        "timestamp_plausible": ts_plausible,
        "producer_known": producer_known,
        "source_path_valid": files_exist,
        "chain_intact": chain_intact,
    }


# ---------------------------------------------------------------- 主流程

def audit_dir(evidence_dir: Path):
    """对单个证据目录做审验，返回结构化结果。"""
    jf, sf, cmds = collect_evidence(evidence_dir)
    chain = chain_check(jf, sf, cmds)
    flags = compute_integrity_flags(chain, cmds, sf)
    integ = evidence_integrity_layer(chain, sf, cmds)
    problems = list(chain["time_consistency"]["problems"])
    if chain["failure_mark_in_text"]:
        problems.append(
            "文本含失败标记字眼（BUILD FAILED/FAILED/Exception…）——可能是描述性提及，需人工复核"
        )
    if not (chain["journal_tool_calls"] or chain["screenshots"] or chain["commands"]):
        problems.append("证据目录为空或无可识别证据文件")
    return {
        "evidence_dir": str(evidence_dir),
        "file_count": {"journal": len(jf), "screenshots": len(sf), "commands": len(cmds)},
        "chain": chain,
        "integrity_flags": flags,
        "evidence_integrity": integ,
        "problems": problems,
        "conclusion": "需人工裁决（机器只出 flag）",
        "integrity_review": {"status": None, "reviewer": None, "reviewed_at": None},
    }


def _find_manifest_in(dirs):
    """在证据目录递归找 evidence_manifest 文件（P2 交付联动）。"""
    for d in dirs:
        for p in sorted(Path(d).rglob("*")):
            if p.is_file() and p.suffix.lower() == ".json" and "manifest" in p.name.lower():
                return p
    return None


def audit_ticket(ticket, table):
    """审验单工单：找证据目录，逐个审验，附加 P0-2 绑定比对 + P2 交付联动。"""
    res = {"ticket": ticket, "audited_at": datetime.datetime.now().isoformat(timespec="seconds")}
    if table is None:
        res["ticket_table"] = "（不可用）"
    else:
        res["ticket_table"] = delivery_binding(ticket, table)
    dirs = find_evidence_dirs(ticket)
    if not dirs:
        res["evidence_dirs"] = []
        res["problems"] = [f"未按工单号 {ticket} 找到证据目录（可能证据在别处/命名不含工单号）"]
        res["integrity_flags"] = {k: {"flag": None, "signal": "无证据可审"} for k in
                                  ("original_rerun", "benchmark_overfit", "semantic_vagueness", "time_inconsistency")}
        res["conclusion"] = "需人工裁决（无证据目录，不判通过/不通过）"
        res["integrity_review"] = {"status": None, "reviewer": None, "reviewed_at": None}
        return res
    # P2 交付联动：工单表已绑定 delivery_id → 证据目录找 manifest 重算比对（不一致标 ⏳失效）
    binding = res.get("ticket_table")
    if isinstance(binding, dict) and binding.get("bound"):
        mf = _find_manifest_in(dirs)
        if mf is None:
            binding["delivery_check"] = {
                "found_manifest": False,
                "note": "已绑定 delivery_id 但证据目录未找到 manifest 文件——无法自动比对，⏳待人工核对交付报告「交付绑定」节",
            }
        else:
            vm = verify_manifest(mf)
            binding["delivery_check"] = {
                "found_manifest": True,
                "manifest": str(mf),
                "match": vm.get("match"),
                "ok": vm.get("ok"),
            }
            if vm.get("match") is False:
                binding["delivery_check"]["flag"] = "⏳失效（manifest_sha 与绑定值不一致——清单内容已变，需新建 delivery_id）"
            elif vm.get("match") is True:
                binding["delivery_check"]["flag"] = "正常（manifest_sha 重算与绑定值一致）"
            else:
                binding["delivery_check"]["flag"] = "待人工核对（manifest_sha 无绑定值）"
    audits = [audit_dir(d) for d in dirs]
    res["evidence_dirs"] = [str(d) for d in dirs]
    res["audits"] = audits
    # 合并 problems
    merged_problems = []
    for a in audits:
        for p in a["problems"]:
            if p not in merged_problems:
                merged_problems.append(p)
    res["problems"] = merged_problems
    res["conclusion"] = "需人工裁决（机器只出 flag）"
    res["integrity_review"] = {"status": None, "reviewer": None, "reviewed_at": None}
    return res


def list_tickets():
    """列出有证据的工单（按证据目录推断 + 工单表对照）。"""
    dirs = []
    for root in EVIDENCE_ROOTS:
        if not root.exists():
            continue
        for p in sorted(root.rglob("*")):
            if p.is_dir():
                # 只列含"工单特征"的目录（含 UPG/S/W 或日期形态），排除深层噪声
                if re.search(r"UPG|S-\d|W-\d|证据|_模拟器|_L2|_复验|_验收|_演示", p.name) and any(
                    p.iterdir() if True else False
                ):
                    dirs.append(p)
    # 简单去重显示
    seen = set()
    lines = []
    for p in dirs:
        key = p.name
        if key in seen:
            continue
        seen.add(key)
        lines.append({"目录": p.name, "路径": str(p.relative_to(ROOT))})
    return lines


# ---------------------------------------------------------------- 输出

def print_text(res):
    if "ticket" in res:
        print(f"═══ 审验 {res['ticket']} ═══")
        if res.get("ticket_table"):
            tt = res["ticket_table"]
            print(f"工单表: {tt.get('note', '')}（delivery_id={tt.get('delivery_id')}）")
            dc = tt.get("delivery_check")
            if isinstance(dc, dict):
                if not dc.get("found_manifest"):
                    print(f"  交付联动: {dc.get('note', '未找到 manifest')}")
                else:
                    print(f"  交付联动: manifest={Path(dc.get('manifest','')).name} match={dc.get('match')} → {dc.get('flag','')}")
        if not res["evidence_dirs"]:
            print(f"证据目录: 未找到")
        else:
            print("证据目录:")
            for d in res["evidence_dirs"]:
                print(f"  - {d}")
        for i, a in enumerate(res.get("audits", []), 1):
            print(f"\n── 证据目录 {i}: {a['evidence_dir']} ──")
            print(f"  文件: journal={a['file_count']['journal']} 截图={a['file_count']['screenshots']} 命令/记录={a['file_count']['commands']}")
            c = a["chain"]
            print(f"  证据链: journal 工具调用 {c['journal_tool_calls']} 次（{c['journal_time_range'][0]} ~ {c['journal_time_range'][1]}）")
            print(f"           时间一致性: {c['time_consistency']['flag']}")
            print("  层A integrity_flags:")
            for k, v in a["integrity_flags"].items():
                mark = "⚠️" if v["flag"] else "·"
                print(f"    {mark} {k}: {v['flag']}  {v['signal'][:80]}")
            print("  层B evidence_integrity:")
            for k, v in a["evidence_integrity"].items():
                print(f"    {k}: {v}")
        print(f"\nproblems: {res['problems']}")
        print(f"conclusion: {res['conclusion']}")
        print(f"integrity_review（人工填）: {res['integrity_review']}")
    else:
        for line in res:
            print(f"{line['目录']}  ←  {line['路径']}")


def main():
    ap = argparse.ArgumentParser(description="SYS-01 P0-3 反伪审验工具（机器只出 flag，人判结论）")
    ap.add_argument("--ticket", help="工单号（如 UPG-28）")
    ap.add_argument("--list", action="store_true", help="列出有证据的工单")
    ap.add_argument("--dir", help="对指定证据目录直接审验")
    ap.add_argument("--manifest", help="独立验证 evidence_manifest 文件（P0-2 绑定）")
    ap.add_argument("--coverage", help="P1-1 能力护栏自动核验（读交付报告「能力护栏」节）")
    ap.add_argument("--ticket-file", help="P1-2 坑位库查询字段核对（读派单文本「已查库」字段）")
    ap.add_argument("--verify-hash", nargs=2, metavar=("branch", "hash"),
                    help="SYS-02 E2 hash 一致性自动闸：校验 <hash> 存在且在 origin/main 祖先（默认主仓库 0027-mov）")
    ap.add_argument("--repo", help="--verify-hash / --verify-hash-self-test 目标 git 仓库（默认主仓库）")
    ap.add_argument("--verify-hash-self-test", action="store_true",
                    help="SYS-02 E2 回归自测：重放 U-49 fixture（9fd39b6→REJECT missing / 2a13dcd→OK）")
    ap.add_argument("--manifest-self-test", action="store_true",
                    help="UPG-86 亲杀锚自测：三坏案（路径嵌注释/缺 sha256/绑定值未写入）全红 + 好案绿")
    ap.add_argument("--json", action="store_true", help="JSON 输出")
    args = ap.parse_args()

    if args.verify_hash_self_test:
        res = verify_hash_self_test(args.repo)
        if args.json:
            print(json_dumps(res))
        else:
            print("═══ E2 回归自测（U-49 fixture 重放）═══")
            print(f"仓库: {res['repo']}")
            for c in res["cases"]:
                mark = "PASS" if c["passed"] else "FAIL"
                print(f"  [{mark}] {c['case']}")
                print(f"        期望 {c['expected']} / 实际 {c['actual']}")
                print(f"        {c['signal']}")
            print(f"结论: {'PASS 2/2' if res['ok'] else 'FAIL'}（机器只出 flag，人裁决）")
        return

    if args.manifest_self_test:
        res = manifest_self_test()
        if args.json:
            print(json_dumps(res))
        else:
            print("═══ UPG-86 manifest 治理自测（三坏案红 + 好案绿）═══")
            for c in res["cases"]:
                mark = "PASS" if c["passed"] else "FAIL"
                print(f"  [{mark}] {c['case']}  (ok={c['actual_ok']})")
                for p in c["problems"][:2]:
                    print(f"        - {p}")
            print(f"结论: {'PASS 4/4' if res['ok'] else 'FAIL'}（机器只出 flag，人裁决）")
        return

    if args.verify_hash:
        branch, h = args.verify_hash
        res = verify_hash(branch, h, args.repo)
        if args.json:
            print(json_dumps(res))
        else:
            print("═══ E2 hash 一致性自动闸 ═══")
            print(f"仓库: {res['repo']}")
            print(f"branch: {res['branch']}  reported_hash: {res['reported_hash']}")
            print(f"结果: {res['status']}" + (f" <{res['reason']}>" if res.get("reason") else ""))
            print(f"signal: {res['signal']}")
            print("结论: 机器只出 flag，登记/放行由人裁决")
        return

    if args.coverage:
        res = verify_coverage(args.coverage)
        if args.json:
            print(json_dumps(res))
        else:
            print(f"能力护栏核验: {res['report']}")
            print(f"coverage_status: {res.get('coverage_status')}")
            flags = res.get("flags", {})
            flag_txt = " + ".join(k for k, v in flags.items() if v) if flags else "（无 flag）"
            print(f"flags: {flag_txt}")
            print(f"signal: {res.get('signal')}")
            if res.get("coverage_decision"):
                print("coverage_decision:", json.dumps(res["coverage_decision"], ensure_ascii=False))
            if res.get("problems"):
                print("problems:")
                for p in res["problems"]:
                    print(f"  - {p}")
            print(f"conclusion: {res['conclusion']}")
        return

    if args.ticket_file:
        res = check_ticket_query_field(args.ticket_file)
        if args.json:
            print(json_dumps(res))
        else:
            print(f"坑位库查询字段核对: {res['ticket_file']}")
            flags = res.get("flags", {})
            flag_txt = " + ".join(k for k, v in flags.items() if v) if flags else "（无 flag）"
            print(f"flags: {flag_txt}")
            print(f"signal: {res.get('signal')}")
            if res.get("problems"):
                print("problems:")
                for p in res["problems"]:
                    print(f"  - {p}")
            print(f"conclusion: {res['conclusion']}")
        return

    if args.manifest:
        res = verify_manifest(args.manifest)
        if args.json:
            print(json_dumps(res))
        else:
            print(f"manifest: {res['manifest']}")
            print(f"delivery_id: {res.get('delivery_id')}  standard_id: {res.get('standard_id')}")
            print(f"manifest_sha 绑定值: {res.get('manifest_sha_bound')}")
            print(f"manifest_sha 重算值: {res.get('manifest_sha_recomputed')}")
            print(f"一致? {res.get('match')}   ok: {res.get('ok')}")
            for e in res.get("evidence_integrity", []):
                print(f"  {e['evidence_id']} path={e['path'][:50]} exists={e['exists']} hash_matches={e['hash_matches']} producer_known={e['producer_known']}")
            if res.get("problems"):
                print("problems:")
                for p in res["problems"]:
                    print(f"  - {p}")
        return

    table, table_err = load_ticket_table()
    if table is None and not (args.list or args.dir):
        print(f"⚠️ {table_err}", file=sys.stderr)

    if args.list:
        lines = list_tickets()
        if args.json:
            print(json_dumps(lines))
        else:
            if not lines:
                print("未发现证据目录")
            else:
                print(f"发现 {len(lines)} 个有证据目录：")
                print_text(lines)
        return

    if args.dir:
        a = audit_dir(Path(args.dir))
        res = {
            "ticket": Path(args.dir).name,
            "ticket_table": {"note": "（--dir 模式，不查工单表）", "delivery_id": None},
            "evidence_dirs": [str(Path(args.dir))],
            "audits": [a],
            "problems": a["problems"],
            "conclusion": a["conclusion"],
            "integrity_review": a["integrity_review"],
        }
    elif args.ticket:
        if table is None and args.ticket:
            # 表不可用时仍可审验证据目录
            res = audit_ticket(args.ticket, None)
        else:
            res = audit_ticket(args.ticket, table)
    else:
        ap.print_help()
        return

    if args.json:
        print(json_dumps(res))
    else:
        print_text(res)


if __name__ == "__main__":
    main()
